import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import db, Publication, PublicationAuthor, PublicationTypeDisplayName, User
from sqlalchemy.orm import joinedload, selectinload # Импортируем для "жадной" загрузки
from datetime import datetime, timedelta # Добавлено timedelta
from typing import Optional     
# --- ОБНОВЛЕННАЯ СИГНАТУРА ФУНКЦИИ ---
def generate_excel_report(user_id: int, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None) -> bytes:
# --- КОНЕЦ ОБНОВЛЕННОЙ СИГНАТУРЫ ---
    """
    Генерирует Excel отчет по публикациям пользователя с измененным заголовком.
    Фильтрует по диапазону дат публикации, если они указаны.
    Жирные границы для строк 1 и 3, тонкие для остальных.

    Args:
        user_id: ID пользователя, для которого генерируется отчет.
        start_date: Начальная дата (включительно) для фильтрации по published_at.
        end_date: Конечная дата (включительно) для фильтрации по published_at.

    Returns:
        Содержимое Excel файла в виде байтов.
    """
    # 1. Загрузка данных пользователя и его публикаций из БД
    # --- ОБНОВЛЕННЫЙ ЗАПРОС К БД ---
    query = Publication.query.filter(
        Publication.user_id == user_id,
        # Раскомментируйте следующую строку, если все еще хотите фильтровать только опубликованные
        Publication.status == 'published'
    ).options(
        selectinload(Publication.authors),
        joinedload(Publication.type),
        joinedload(Publication.display_name)
    )

    # Применяем фильтры по датам, если они заданы
    if start_date:
        # Убедимся, что published_at существует и больше или равно начальной дате
        query = query.filter(Publication.published_at != None, Publication.published_at >= start_date)
        print(f"Фильтр: >= {start_date}") # Отладка

    if end_date:
        # Для включения публикаций за *весь* конечный день,
        # нужно фильтровать по началу *следующего* дня
        end_of_day_inclusive = end_date + timedelta(days=1)
        query = query.filter(Publication.published_at != None, Publication.published_at < end_of_day_inclusive)
        print(f"Фильтр: < {end_of_day_inclusive}") # Отладка

    # Сортировка остается прежней
    publications = query.order_by(
        Publication.year.desc(), Publication.title
    ).all()
    # --- КОНЕЦ ОБНОВЛЕННОГО ЗАПРОСА К БД ---

    print(f"Найдено публикаций для отчета ({user_id}, {start_date}-{end_date}): {len(publications)}") # Отладка

    # 2. Создание Excel Workbook и Worksheet
    # ... (остальной код функции без изменений) ...
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Публикации"

    # --- Определение стилей ---
    base_font_size = 9
    header_font = Font(bold=True, size=base_font_size)
    data_font = Font(size=base_font_size)
    bold_data_font = Font(bold=True, size=base_font_size) # Для K2/L2
    italic_font = Font(italic=True, size=base_font_size)

    # Выравнивания
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    number_row_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(vertical='top', wrap_text=True)

    # Стили границ
    thin_border = Border(left=Side(style='thin', color='000000'),
                         right=Side(style='thin', color='000000'),
                         top=Side(style='thin', color='000000'),
                         bottom=Side(style='thin', color='000000'))
    medium_border = Border(left=Side(style='medium', color='000000'),
                           right=Side(style='medium', color='000000'),
                           top=Side(style='medium', color='000000'),
                           bottom=Side(style='medium', color='000000'))

    # 3. Определение заголовков
    headers_original_texts = [
        "№\nп/п", "Автор (ФИО)", "Наименование работы",
        "Наименование журнала/конференции/симпозиума", "DOI", "ISSN", "ISBN", "Q",
        "Номер, том, выпуск, страницы", "Статус\n1(WoS)\n1(Scop)\n2(BAK)",
        "Количество\nавторов\n(всего)", "в т.ч.\nработн\nиков\nвуза", "Кафедра",
        "ФИО\nсоавторов", "Вид работы", "Издательство", "Место изд-ва",
        "Год\nиздания", "Объем в п.л.", "Тираж",
        "Направление и код по\nклассификатору", "Примечание"
    ]
    moved_header_k_text = headers_original_texts[10]
    moved_header_l_text = headers_original_texts[11]
    num_columns = len(headers_original_texts)

    # 4. Заполнение строки заголовков (Строка 1) - ЖИРНЫЕ границы
    for col_num in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col_num)
        if col_num != 11 and col_num != 12:
            cell.value = headers_original_texts[col_num-1]
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = medium_border # <--- Жирные границы
    ws.row_dimensions[1].height = 156 # Высота

    # Объединение K1:L1 и стили
    ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=12)
    merged_cell_k1l1 = ws.cell(row=1, column=11)
    merged_cell_k1l1.value = "Количество авторов"
    merged_cell_k1l1.alignment = header_alignment
    merged_cell_k1l1.border = medium_border # <--- Жирные границы

    # 5. Вставка ПУСТОЙ строки (новая Строка 2) - ТОНКИЕ границы
    ws.insert_rows(2)
    for col_num in range(1, num_columns + 1):
         empty_cell = ws.cell(row=2, column=col_num)
         empty_cell.border = thin_border # <--- Тонкие границы

    # 6. Заполнение ячеек K2 и L2 (Строка 2) - ТОНКИЕ границы
    cell_k2 = ws.cell(row=2, column=11)
    cell_k2.value = moved_header_k_text
    cell_k2.font = bold_data_font # Жирный шрифт
    cell_k2.alignment = header_alignment
    cell_k2.border = thin_border # <--- Тонкие границы

    cell_l2 = ws.cell(row=2, column=12)
    cell_l2.value = moved_header_l_text
    cell_l2.font = bold_data_font # Жирный шрифт
    cell_l2.alignment = header_alignment
    cell_l2.border = thin_border # <--- Тонкие границы

    # 7. Заполнение строки нумерации (новая Строка 3) - ЖИРНЫЕ границы
    ws.insert_rows(3)
    for col_num in range(1, num_columns + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = col_num
        cell.font = data_font
        cell.alignment = number_row_alignment
        cell.border = medium_border # <--- Жирные границы
    # ws.row_dimensions[3].height = 15

    # 8. Заполнение данными (Начинаем с row_index = 4) - ТОНКИЕ границы
    current_data_row_index = 4

    if not publications:
        # Сообщение "Публикации не найдены"
        ws.merge_cells(start_row=current_data_row_index, start_column=1, end_row=current_data_row_index, end_column=num_columns)
        merged_cell = ws.cell(row=current_data_row_index, column=1)
        # --- ОБНОВЛЕНО СООБЩЕНИЕ ---
        message = "Публикации не найдены"
        if start_date or end_date:
            message += f" за период"
            if start_date: message += f" с {start_date.strftime('%d.%m.%Y')}"
            if end_date: message += f" по {end_date.strftime('%d.%m.%Y')}"
        merged_cell.value = message + "."
        # --- КОНЕЦ ОБНОВЛЕНИЯ СООБЩЕНИЯ ---

        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.font = italic_font
        merged_cell.border = thin_border # <--- Тонкие границы
    else:
        for i, pub in enumerate(publications, 1):
            # --- Логика извлечения и форматирования данных --- (без изменений)
            first_author_name = pub.authors[0].name if pub.authors else ""
            num_vol_pages_parts = []
            if pub.volume: num_vol_pages_parts.append(f"Т. {pub.volume}")
            if pub.number: num_vol_pages_parts.append(f"№ {pub.number}")
            if pub.pages:
                pages_formatted = str(pub.pages).replace('-', '–')
                num_vol_pages_parts.append(f"С. {pages_formatted}")
            num_vol_pages_str = ", ".join(num_vol_pages_parts)
            status_parts = []
            if pub.is_wos: status_parts.append("1(WoS)")
            if pub.is_scopus: status_parts.append("1(Scop)")
            if pub.is_vak: status_parts.append("2(BAK)")
            status_str = " ".join(status_parts)
            total_authors = len(pub.authors)
            employee_authors = sum(1 for author in pub.authors if author.is_employee)
            coauthors = ", ".join(author.name for author in pub.authors[1:]) if total_authors > 1 else ""

            work_type = ""
            if pub.display_name and pub.display_name.display_name == "Конференция РИНЦ":
                work_type = "Конференция РИНЦ"
            elif pub.type and pub.type.name:
                type_name_lower = pub.type.name.lower()
                if type_name_lower == 'article': work_type = "Статья"
                elif type_name_lower == 'conference': work_type = "Конференция"
                else: work_type = pub.display_name.display_name if pub.display_name else pub.type.name
            elif pub.display_name and pub.display_name.display_name:
                work_type = pub.display_name.display_name

            data_row_values = [
                i, first_author_name, pub.title or "",
                pub.journal_conference_name or "", pub.doi or "", pub.issn or "",
                pub.isbn or "", pub.quartile or "", num_vol_pages_str, status_str,
                total_authors, employee_authors, pub.department or "", coauthors,
                work_type, pub.publisher or "", pub.publisher_location or "", pub.year or "",
                pub.printed_sheets_volume if pub.printed_sheets_volume is not None else "",
                pub.circulation if pub.circulation is not None else "",
                pub.classification_code or "", pub.notes or ""
            ]

            # --- Заполнение ячеек данными и применение стилей ---
            for col_idx, value in enumerate(data_row_values, 1):
                cell = ws.cell(row=current_data_row_index, column=col_idx)
                cell.value = value
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
            current_data_row_index += 1

    # 9. Настройка ширины колонок (без изменений)
    initial_widths = {
        'A': 3.3, 'B': 18.57, 'C': 23.29, 'D': 23.29, 'E': 9.29, 'F': 7.57,
        'G': 7.57, 'H': 6.86, 'I': 9.86, 'J': 6, 'K': 5.57, 'L': 4.86,
        'M': 8.86, 'N': 10.57, 'O': 12.43, 'P': 15.86, 'Q': 7.43, 'R': 8.71,
        'S': 6.71, 'T': 6.14, 'U': 16.57, 'V': 10.71
    }
    width_increment = 0.71
    for col_idx, col_letter in enumerate(map(get_column_letter, range(1, num_columns + 1)), 1):
        initial_width = initial_widths.get(col_letter, 10)
        adjusted_width = initial_width + width_increment
        ws.column_dimensions[col_letter].width = adjusted_width

    # 10. Сохранение в байтовый поток
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()